#!/usr/bin/env python3
"""
sync-catalog.py
Imports Abel_Catalog_CLEAN.xlsx → Prisma Product table + BOM entries.

Usage:
  python3 scripts/sync-catalog.py                # Full sync
  python3 scripts/sync-catalog.py --dry-run      # Preview only
  python3 scripts/sync-catalog.py --skip-bom     # Skip BOM import
  python3 scripts/sync-catalog.py --bom-only     # Only BOM import
"""

import os, sys, re, json
from datetime import datetime

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from urllib.parse import urlparse

# ── Config ──────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
CATALOG_PATH = os.path.join(os.path.dirname(PROJECT_DIR), "Abel_Catalog_CLEAN.xlsx")

DRY_RUN = "--dry-run" in sys.argv
BOM_ONLY = "--bom-only" in sys.argv
SKIP_BOM = "--skip-bom" in sys.argv


def get_db_url():
    """Read DATABASE_URL from .env file."""
    env_path = os.path.join(PROJECT_DIR, ".env")
    if not os.path.exists(env_path):
        env_path = os.path.join(PROJECT_DIR, ".env.local")
    if not os.path.exists(env_path):
        print("❌ No .env or .env.local found")
        sys.exit(1)

    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line.startswith("DATABASE_URL="):
                url = line.split("=", 1)[1].strip().strip('"').strip("'")
                return url

    print("❌ DATABASE_URL not found in .env")
    sys.exit(1)


def connect_db():
    url = get_db_url()
    parsed = urlparse(url)
    conn = psycopg2.connect(
        host=parsed.hostname,
        port=parsed.port or 5432,
        dbname=parsed.path.lstrip("/").split("?")[0],
        user=parsed.username,
        password=parsed.password,
        sslmode="require",
    )
    conn.autocommit = False
    return conn


# ── Normalization helpers ───────────────────────────────────────────
def clean(val):
    if val is None or str(val).strip() == "":
        return None
    return str(val).strip()


def clean_num(val):
    if val is None or str(val).strip() == "":
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def map_category(clean_cat):
    if not clean_cat:
        return "Other"
    cat = clean_cat.lower()
    if "exterior door" in cat:
        return "Exterior Doors"
    if any(x in cat for x in ["interior door", "hollow core", "solid core", "bifold", "pocket", "barn door"]):
        return "Interior Doors"
    if "fire-rated" in cat:
        return "Fire-Rated Doors"
    if "patio" in cat or "sliding glass" in cat:
        return "Patio Doors"
    if "garage" in cat:
        return "Exterior Doors"
    if "door frame" in cat:
        return "Door Frames"
    if "door slab" in cat:
        return "Door Slabs"
    if "trim" in cat or "molding" in cat or "moulding" in cat:
        return "Trim & Molding"
    if "hardware" in cat:
        return "Hardware"
    if "jamb" in cat:
        return "Jambs"
    if "stair" in cat:
        return "Stair Parts"
    if "closet" in cat or "shelf" in cat:
        return "Closet & Shelf"
    if "lumber" in cat or "sheet good" in cat:
        return "Lumber & Sheet Goods"
    if "glass" in cat or "lite" in cat:
        return "Glass & Inserts"
    if "threshold" in cat:
        return "Thresholds"
    if "weather" in cat:
        return "Weatherstripping"
    if "attic" in cat:
        return "Attic Access"
    if "hvac" in cat:
        return "HVAC Doors"
    if "service" in cat or "labor" in cat:
        return "Services & Labor"
    if "building material" in cat:
        return "Building Materials"
    if "window" in cat:
        return "Window Components"
    if "dunnage" in cat:
        return "Dunnage"
    return clean_cat


def map_subcategory(clean_cat):
    if not clean_cat:
        return None
    parts = clean_cat.split(" - ")
    return parts[1].strip() if len(parts) > 1 else None


def normalize_door_size(raw):
    if not raw:
        return None
    m = re.search(r'(\d+)["\u2033]?\s*x\s*(\d+)', str(raw))
    if m:
        return f"{m.group(1)}{m.group(2)}"
    return str(raw).strip()


def normalize_hardware_finish(raw):
    if not raw:
        return None
    f = str(raw).lower().strip()
    mapping = {
        "satin nickel": "SN", "sn": "SN",
        "oil rubbed bronze": "ORB", "orb": "ORB",
        "black": "BLK", "blk": "BLK", "matte black": "BLK",
        "antique brass": "AB", "ab": "AB",
        "satin chrome": "SC", "sc": "SC",
        "polished chrome": "PC", "pc": "PC",
    }
    return mapping.get(f, str(raw).strip())


def normalize_casing(raw):
    if not raw:
        return None
    c = str(raw).lower().strip()
    if "a-col" in c or "a-colonial" in c or "2-1/4" in c:
        return "A-Col"
    if "colonial" in c or "3-1/4" in c or "c-322" in c:
        return "C-322"
    if "no casing" in c or "none" in c:
        return None
    return str(raw).strip()


def gen_cuid():
    """Generate a cuid-like ID."""
    import random, time
    ts = hex(int(time.time() * 1000))[2:]
    rand = "".join(random.choices("abcdefghijklmnopqrstuvwxyz0123456789", k=12))
    return f"cl{ts}{rand}"


# ── Product Sync ────────────────────────────────────────────────────
def sync_products(cur, rows):
    print(f"\n📦 Syncing {len(rows)} products...")

    upserted = 0
    skipped = 0
    errors = 0

    for i, row in enumerate(rows):
        try:
            sku = clean(row.get("SKU"))
            if not sku:
                skipped += 1
                continue

            name = clean(row.get("Product Name")) or sku
            clean_cat = clean(row.get("Clean Category")) or "Other"
            cost = clean_num(row.get("Unit Cost"))
            list_price = clean_num(row.get("Default List Price"))
            base_price = list_price if list_price > 0 else cost * 1.35
            margin = (list_price - cost) / list_price if list_price > 0 else 0.25
            min_margin = max(0.10, min(margin, 0.60))
            is_active = row.get("Is Active") is not False and row.get("Is Active") is not None

            category = map_category(clean_cat)
            subcategory = map_subcategory(clean_cat) or clean(row.get("Product Type"))
            door_size = normalize_door_size(clean(row.get("Door Size")))
            handing = clean(row.get("Handing"))
            core_type = clean(row.get("Core Type"))
            panel_style = clean(row.get("Panel Style"))
            jamb_size = clean(row.get("Jamb Size"))
            casing_code = normalize_casing(clean(row.get("Casing")))
            hardware_finish = normalize_hardware_finish(clean(row.get("Hardware Finish")))
            material = clean(row.get("Material"))
            fire_rating = clean(row.get("Fire Rating"))
            inflow_cat = clean(row.get("Original InFlow Category"))

            if DRY_RUN:
                upserted += 1
                continue

            # Upsert: INSERT ... ON CONFLICT(sku) UPDATE
            cur.execute(
                """
                INSERT INTO "Product" (
                    id, sku, name, "displayName", category, subcategory,
                    cost, "basePrice", "minMargin",
                    "doorSize", handing, "coreType", "panelStyle",
                    "jambSize", "casingCode", "hardwareFinish", material, "fireRating",
                    active, "inStock", "inflowCategory", "lastSyncedAt",
                    "createdAt", "updatedAt"
                ) VALUES (
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, true, %s, NOW(),
                    NOW(), NOW()
                )
                ON CONFLICT (sku) DO UPDATE SET
                    name = EXCLUDED.name,
                    "displayName" = EXCLUDED."displayName",
                    category = EXCLUDED.category,
                    subcategory = EXCLUDED.subcategory,
                    cost = EXCLUDED.cost,
                    "basePrice" = EXCLUDED."basePrice",
                    "minMargin" = EXCLUDED."minMargin",
                    "doorSize" = EXCLUDED."doorSize",
                    handing = EXCLUDED.handing,
                    "coreType" = EXCLUDED."coreType",
                    "panelStyle" = EXCLUDED."panelStyle",
                    "jambSize" = EXCLUDED."jambSize",
                    "casingCode" = EXCLUDED."casingCode",
                    "hardwareFinish" = EXCLUDED."hardwareFinish",
                    material = EXCLUDED.material,
                    "fireRating" = EXCLUDED."fireRating",
                    active = EXCLUDED.active,
                    "inflowCategory" = EXCLUDED."inflowCategory",
                    "lastSyncedAt" = NOW(),
                    "updatedAt" = NOW()
                """,
                (
                    gen_cuid(), sku, name, name, category, subcategory,
                    cost, base_price, min_margin,
                    door_size, handing, core_type, panel_style,
                    jamb_size, casing_code, hardware_finish, material, fire_rating,
                    is_active, inflow_cat,
                ),
            )
            upserted += 1

        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"  ❌ Error on SKU '{row.get('SKU')}': {e}")

        # Progress
        if (i + 1) % 200 == 0 or i == len(rows) - 1:
            pct = round((i + 1) / len(rows) * 100)
            print(f"\r  Progress: {pct}% ({i + 1}/{len(rows)})", end="", flush=True)

    print(f"\n  ✅ Products: {upserted} upserted, {skipped} skipped, {errors} errors")
    return upserted, skipped, errors


# ── BOM Sync ────────────────────────────────────────────────────────
def sync_bom(cur, rows):
    print(f"\n🔧 Syncing {len(rows)} BOM entries...")

    # Build SKU → ID and Name → ID lookup
    cur.execute('SELECT id, sku, name FROM "Product" WHERE active = true')
    products = cur.fetchall()
    sku_to_id = {r[1]: r[0] for r in products}
    name_to_id = {r[2].lower(): r[0] for r in products}

    # Clear existing BOMs
    if not DRY_RUN:
        cur.execute('DELETE FROM "BomEntry"')
        print(f"  🗑️  Cleared existing BOM entries")

    created = 0
    skipped = 0
    errors = 0

    for i, row in enumerate(rows):
        try:
            parent_sku = clean(row.get("Finished Product SKU"))
            component_name = clean(row.get("Component Name"))
            if not parent_sku or not component_name:
                skipped += 1
                continue

            parent_id = sku_to_id.get(parent_sku)
            if not parent_id:
                skipped += 1
                continue

            component_id = name_to_id.get(component_name.lower())
            if not component_id:
                skipped += 1
                continue

            if parent_id == component_id:
                skipped += 1  # Can't be own component
                continue

            if DRY_RUN:
                created += 1
                continue

            qty = clean_num(row.get("Quantity")) or 1.0
            uom = clean(row.get("UOM")) or "ea"

            cur.execute(
                """INSERT INTO "BomEntry" (id, "parentProductId", "componentProductId", quantity, unit, "createdAt")
                   VALUES (%s, %s, %s, %s, %s, NOW())""",
                (gen_cuid(), parent_id, component_id, qty, uom),
            )
            created += 1

        except Exception as e:
            errors += 1
            if errors <= 3:
                print(f"  ❌ BOM error: {e}")

        if (i + 1) % 500 == 0 or i == len(rows) - 1:
            pct = round((i + 1) / len(rows) * 100)
            print(f"\r  Progress: {pct}% ({i + 1}/{len(rows)})", end="", flush=True)

    print(f"\n  ✅ BOM: {created} created, {skipped} skipped, {errors} errors")
    return created, skipped, errors


# ── Validation ──────────────────────────────────────────────────────
def validate(cur):
    print("\n📊 Validating data integrity...")

    cur.execute('SELECT category, COUNT(*) as cnt FROM "Product" WHERE active = true GROUP BY category ORDER BY cnt DESC')
    cats = cur.fetchall()
    print("\n  Categories:")
    for cat, cnt in cats:
        print(f"    {cat}: {cnt}")

    cur.execute('SELECT COUNT(*) FROM "Product"')
    total = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "Product" WHERE active = true')
    active = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "Product" WHERE cost > 0')
    with_cost = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "Product" WHERE "basePrice" > 0')
    with_price = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "BomEntry"')
    bom_count = cur.fetchone()[0]

    print(f"\n  📈 Summary:")
    print(f"    Total products:  {total}")
    print(f"    Active:          {active}")
    print(f"    With cost > 0:   {with_cost}")
    print(f"    With price > 0:  {with_price}")
    print(f"    BOM entries:     {bom_count}")

    # Attribute coverage
    attrs = ["doorSize", "handing", "coreType", "panelStyle", "hardwareFinish", "material", "jambSize", "casingCode", "fireRating"]
    print(f"\n  🏷️  Attribute Coverage:")
    for attr in attrs:
        cur.execute(f'SELECT COUNT(*) FROM "Product" WHERE "{attr}" IS NOT NULL')
        cnt = cur.fetchone()[0]
        pct = round(cnt / total * 100) if total > 0 else 0
        print(f"    {attr}: {cnt} ({pct}%)")


# ── Main ────────────────────────────────────────────────────────────
def main():
    print("═" * 55)
    print("  Abel Lumber Catalog Sync")
    print(f"  Source: {CATALOG_PATH}")
    print(f"  Mode: {'🔍 DRY RUN' if DRY_RUN else '🚀 LIVE'}")
    print("═" * 55)

    if not os.path.exists(CATALOG_PATH):
        print(f"❌ File not found: {CATALOG_PATH}")
        sys.exit(1)

    # Load XLSX
    wb = openpyxl.load_workbook(CATALOG_PATH, read_only=True, data_only=True)
    print(f"\n📖 Sheets: {wb.sheetnames}")

    def sheet_to_dicts(sheet_name):
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(next(rows_iter))]
        result = []
        for row in rows_iter:
            if all(v is None for v in row):
                continue
            result.append(dict(zip(headers, row)))
        return result

    conn = connect_db()
    cur = conn.cursor()

    try:
        # ── Products ──
        if not BOM_ONLY:
            product_rows = sheet_to_dicts(wb.sheetnames[0])
            print(f"   Product Master: {len(product_rows)} rows")
            sync_products(cur, product_rows)
            if not DRY_RUN:
                conn.commit()

        # ── BOM ──
        if not SKIP_BOM:
            bom_sheet = next((s for s in wb.sheetnames if "bom" in s.lower()), None)
            if bom_sheet:
                bom_rows = sheet_to_dicts(bom_sheet)
                print(f"   BOM Explorer: {len(bom_rows)} rows")
                sync_bom(cur, bom_rows)
                if not DRY_RUN:
                    conn.commit()

        # ── Validate ──
        if not DRY_RUN:
            validate(cur)

    except Exception as e:
        conn.rollback()
        print(f"\n❌ Fatal error: {e}")
        raise
    finally:
        cur.close()
        conn.close()
        wb.close()

    print("\n" + "═" * 55)
    print(f"  {'✅ Dry run complete — no data written' if DRY_RUN else '✅ Catalog sync complete!'}")
    print("═" * 55 + "\n")


if __name__ == "__main__":
    main()
